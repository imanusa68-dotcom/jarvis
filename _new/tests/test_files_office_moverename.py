"""
Tests for the three follow-up fixes:
  - Problem 2: real Office documents (.docx / .xlsx), not renamed text.
  - Problem 3: move / rename unblocked behind confirm, boundary-checked.

All filesystem tests redirect safe-roots to a temp dir — the real home is
never touched.

Run:  python -m pytest tests/test_files_office_moverename.py -q
"""

import tempfile
from pathlib import Path

import actions.file_controller as fc
from core.security import check_tool_call, needs_confirmation


class _TempRoots:
    def __enter__(self):
        self._tmp = tempfile.TemporaryDirectory()
        root = Path(self._tmp.name) / "Desktop"
        root.mkdir(parents=True, exist_ok=True)
        self._orig = fc._safe_roots
        fc._safe_roots = lambda: [root]  # type: ignore
        return root

    def __exit__(self, *exc):
        fc._safe_roots = self._orig  # type: ignore
        self._tmp.cleanup()


# ── Problem 2: real Office documents ────────────────────────────────────────

def test_docx_is_a_real_word_document():
    with _TempRoots() as root:
        target = root / "дипломная работа.docx"
        r = fc.create_file(str(target), content="Заголовок\nАбзац")
        assert target.exists() and target.stat().st_size > 1000  # real zip, not 0
        from docx import Document
        doc = Document(str(target))
        assert [p.text for p in doc.paragraphs] == ["Заголовок", "Абзац"]
        assert "disabled" not in r.lower()


def test_xlsx_is_a_real_spreadsheet():
    with _TempRoots() as root:
        target = root / "таблица.xlsx"
        fc.create_file(str(target), content="a,b\n1,2")
        from openpyxl import load_workbook
        ws = load_workbook(str(target)).active
        assert ws["A1"].value == "a" and ws["B2"].value == "2"


def test_office_still_boundary_guarded():
    with _TempRoots():
        import os
        outside = "C:/Windows/x.docx" if os.name == "nt" else "/etc/x.docx"
        r = fc.create_file(outside, content="x")
        assert "SECURITY" in r


# ── Problem 3: move / rename ────────────────────────────────────────────────

def test_move_actually_moves_within_safe_roots():
    with _TempRoots() as root:
        src = root / "a.txt"
        fc.create_file(str(src), content="x")
        (root / "sub").mkdir()
        r = fc.move_file(str(src), str(root / "sub"))
        assert not src.exists()
        assert (root / "sub" / "a.txt").exists()
        assert "moved" in r.lower()


def test_rename_actually_renames():
    with _TempRoots() as root:
        src = root / "old.txt"
        fc.create_file(str(src), content="x")
        r = fc.rename_file(str(src), "new.txt")
        assert not src.exists() and (root / "new.txt").exists()
        assert "renamed" in r.lower()


def test_rename_strips_path_components():
    # new_name must be a bare name — never let it escape to another folder.
    with _TempRoots() as root:
        src = root / "f.txt"
        fc.create_file(str(src), content="x")
        fc.rename_file(str(src), "../../evil.txt")
        # The escape target must NOT exist; a bare 'evil.txt' stays in-folder.
        assert not (root / ".." / ".." / "evil.txt").resolve().exists()


def test_move_refuses_destination_outside_safe_roots():
    with _TempRoots() as root:
        src = root / "a.txt"
        fc.create_file(str(src), content="x")
        import os
        dst = "C:/Windows/a.txt" if os.name == "nt" else "/etc/a.txt"
        r = fc.move_file(str(src), dst)
        assert "SECURITY" in r
        assert src.exists()  # not moved


def test_move_does_not_overwrite_existing():
    with _TempRoots() as root:
        a = root / "a.txt"; b = root / "b.txt"
        fc.create_file(str(a), content="A")
        fc.create_file(str(b), content="B")
        r = fc.move_file(str(a), str(b))
        assert "already exists" in r.lower()
        assert a.exists() and b.read_text(encoding="utf-8") == "B"


def test_policy_move_rename_allowed_and_confirm():
    for action in ("move", "rename"):
        d = check_tool_call("file_controller", {"action": action})
        assert d.allowed
        assert needs_confirmation("file_controller", {"action": action})[0] is True


def test_policy_delete_allowed_behind_confirmation():
    # Stage 2.6: undoable delete (staged copy + Recycle Bin) behind a confirm.
    d = check_tool_call("file_controller", {"action": "delete", "path": "desktop"})
    assert d.allowed
    assert d.policy == "confirm", d.policy


if __name__ == "__main__":
    import sys
    tests = [
        test_docx_is_a_real_word_document,
        test_xlsx_is_a_real_spreadsheet,
        test_office_still_boundary_guarded,
        test_move_actually_moves_within_safe_roots,
        test_rename_actually_renames,
        test_rename_strips_path_components,
        test_move_refuses_destination_outside_safe_roots,
        test_move_does_not_overwrite_existing,
        test_policy_move_rename_allowed_and_confirm,
        test_policy_delete_allowed_behind_confirmation,
    ]
    fails = 0
    for fn in tests:
        try:
            fn()
            print(f"  OK   {fn.__name__}")
        except AssertionError as e:
            fails += 1
            print(f"  FAIL {fn.__name__}: {e}")
        except Exception as e:
            fails += 1
            print(f"  ERR  {fn.__name__}: {e!r}")
    print("\nRESULT:", "ALL PASS" if fails == 0 else f"{fails} FAILURES")
    sys.exit(1 if fails else 0)
